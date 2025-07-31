import argparse
import json
from utils.video import extract_frames
from utils.blip import caption_image
from utils.similarity import caption_similarity, relevance_to_goal
import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage
import io

def run_ingestion(video_path, goal=None, interval=5, similarity_threshold=0.75):
    frames = extract_frames(video_path, interval_sec=interval)
    captions = [caption_image(path) for _, path in frames]

    segments = []
    prev_caption = captions[0]
    start_time = frames[0][0]

    for i in range(1, len(frames)):
        current_time, frame_path = frames[i]
        similarity = caption_similarity(prev_caption, captions[i])
        
        if similarity < similarity_threshold:
            segment = {
                "start": start_time,
                "end": current_time,
                "timestamp": seconds_to_timestamp(start_time),
                "caption": prev_caption,
                "key_frame": frames[i-1][1]
            }
            if goal:
                segment["relevance"] = relevance_to_goal(prev_caption, goal)
            segments.append(segment)
            start_time = current_time
            prev_caption = captions[i]
    
    # Add final segment
    segments.append({
        "start": start_time,
        "end": frames[-1][0],
        "timestamp": seconds_to_timestamp(start_time),
        "caption": captions[-1],
        "key_frame": frames[-1][1],
        "relevance": relevance_to_goal(captions[-1], goal) if goal else None
    })

    # Optional filtering
    if goal:
        segments = [s for s in segments if s["relevance"] > 0.5]

    with open("metadata.json", "w") as f:
        json.dump(segments, f, indent=2)
    export_segments_to_excel_with_images(segments)
    print(f"[📄] Also saved scene metadata to scene_metadata.xlsx")
    
    print(f"[✔] Saved {len(segments)} segments to metadata.json")
def seconds_to_timestamp(seconds):
    hrs = seconds // 3600
    mins = (seconds % 3600) // 60
    secs = seconds % 60
    return f"{int(hrs):02}:{int(mins):02}:{int(secs):02}"
def export_segments_to_excel_with_images(segments, output_path="scene_metadata.xlsx"):
    df = pd.DataFrame(segments)
    df.to_excel(output_path, index=False)

    wb = load_workbook(output_path)
    ws = wb.active

    key_col_name = "key_frame"
    if key_col_name not in df.columns:
        print("❌ No 'key_frame' column found in segments. Cannot embed images.")
        return

    key_col_index = df.columns.get_loc(key_col_name) + 1
    col_letter = get_column_letter(key_col_index)
    ws.column_dimensions[col_letter].width = 25  # adjust if needed

    for i, segment in enumerate(segments):
        img_path = segment.get("key_frame")
        row_num = i + 2  # Excel rows are 1-indexed

        if not img_path:
            continue

        try:
            pil_img = PILImage.open(img_path)
            pil_img.thumbnail((160, 90))  # Resize to fit cell nicely

            # Save to memory
            img_bytes = io.BytesIO()
            pil_img.save(img_bytes, format="PNG")
            img_bytes.seek(0)
            img = ExcelImage(img_bytes)

            # Insert image
            cell = f"{col_letter}{row_num}"
            ws.add_image(img, cell)

            # 🔥 Match row height to image height (Excel ≈ pixels * 0.75)
            ws.row_dimensions[row_num].height = img.height * 0.75

        except Exception as e:
            print(f"⚠️ Could not embed image at row {row_num}: {e}")

    wb.save(output_path)
    print(f"[🖼️] Excel file saved with embedded thumbnails → {output_path}")

if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--video_path", required=True)
    parser.add_argument("--goal", help="Optional prompt to focus on")
    parser.add_argument("--frame_interval", type=int, default=5)
    args = parser.parse_args()

    run_ingestion(args.video_path, goal=args.goal, interval=args.frame_interval)
